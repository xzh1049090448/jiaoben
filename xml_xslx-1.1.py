from xml.dom.minidom import parse
import xml.dom.minidom
from matplotlib.pyplot import title
import openpyxl

xml_file = "info.xml"
txt_files = ["101001.txt"]
output_file_name = 'biaoge.xlsx'
tag_list = ["条件一","条件二","条件三","条件四"]
tag_str = "Y"
titles = ["InfoId","InfoName","SetName","条件一","条件二","条件三","条件四","Path"]
tag_mp = {}
def get_xml_info(xml_file):
    DOMTree = xml.dom.minidom.parse(xml_file)
    collection = DOMTree.documentElement
    setinfos = collection.getElementsByTagName("SetInfo")
    res = {}
    for setinfo in setinfos:
        setName = setinfo.getAttribute("name")
        info_list = setinfo.getElementsByTagName('InfoItem')
        for info in info_list:
            info_id = info.getAttribute("infoId")
            res[info_id] = {}
            res[info_id]["SetName"] = setName
            res[info_id]["InfoName"] = info.getElementsByTagName('name')[0].childNodes[0].data
            path1_els = info.getElementsByTagName('path1')
            path2_els = info.getElementsByTagName('path2')
            if path1_els and path2_els:
                res[info_id]["Path"] = path1_els[0].childNodes[0].data+'/'+path2_els[0].childNodes[0].data
    return res

def save_excel(data_list, xlsx_file):
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb['Sheet1']
    rows = len(data_list)
    local_row = ws.max_row
    tmp_info = []
    for i in range(1,local_row+1):
        info_id = ws.cell(row=i+1,column=1).value
        if info_id in tag_mp:
            tmp_info.append(info_id)
            for tag in tag_mp[info_id]:
                ws.cell(row=i+1,column=titles.index(tag)+1).value = tag_str
    for i in range(rows):
        for j in range(len(titles)):
            if data_list[i].get("InfoId") not in tmp_info:
                ws.cell(row=i + 1 + local_row, column=j + 1).value = data_list[i].get(titles[j],"")
    wb.save(filename=xlsx_file)


info_all = get_xml_info(xml_file)
error_list = []
data_list = []
for i in range(len(txt_files)):
    with open(txt_files[i],'r') as fw:
        for info_id in fw.readlines():
            info_id = info_id.strip()
            if info_id not in info_all:
                error_list.append(info_id)
                continue
            if info_id in tag_mp:
                tag_mp[info_id].append(tag_list[i])
                for info in data_list:
                    if info["InfoId"] == info_id:
                        info["InfoId"][tag_list[i]] = tag_str
            else:
                tag_mp[info_id] = [tag_list[i]]
                mp = {}
                mp["InfoId"] = info_id
                mp[tag_list[i]] = tag_str
                for key in info_all[info_id]:
                    mp[key] = info_all[info_id][key]
                data_list.append(mp)
print(tag_mp)
print(data_list)
for info_id in error_list:
    print(f"Error {info_id} not exist!")
save_excel(data_list, output_file_name)